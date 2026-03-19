import streamlit as st
import pandas as pd
import requests
import base64
import json
import uuid
import io
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go

st.set_page_config(
    page_title="Roesel Transportes",
    page_icon="🚛",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── CSS PROFISSIONAL (ESTILO DASHBOARD) ────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

* { font-family: 'Inter', sans-serif !important; box-sizing: border-box; }

/* Fundo principal */
.stApp { background: #F4F6F9 !important; }

/* ── SIDEBAR ── */
[data-testid="stSidebar"] {
    background: #1A3A5C !important;
    border-right: none !important;
    min-width: 220px !important;
    max-width: 220px !important;
}
[data-testid="stSidebar"] * { color: #B8CDE0 !important; }
[data-testid="stSidebar"] .stRadio label {
    padding: 10px 16px !important;
    border-radius: 8px !important;
    display: block !important;
    font-size: 14px !important;
    font-weight: 500 !important;
    cursor: pointer !important;
    transition: all 0.15s !important;
    color: #B8CDE0 !important;
}
[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(255,255,255,0.08) !important;
    color: white !important;
}
[data-testid="stSidebar"] [data-baseweb="radio"] input:checked + div + label,
[data-testid="stSidebar"] [data-baseweb="radio"] input:checked ~ label {
    background: rgba(255,255,255,0.15) !important;
    color: white !important;
}
[data-testid="stSidebar"] .stSelectbox > div > div {
    background: rgba(255,255,255,0.1) !important;
    border: 1px solid rgba(255,255,255,0.2) !important;
    color: white !important;
}
[data-testid="stSidebar"] .stButton > button {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(255,255,255,0.15) !important;
    color: #B8CDE0 !important;
    font-size: 13px !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(255,255,255,0.15) !important;
    color: white !important;
}
[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.15) !important; }

/* ── TOPBAR ── */
[data-testid="stHeader"] { background: white !important; border-bottom: 1px solid #E5E9F0; }

/* ── CONTEÚDO PRINCIPAL ── */
.main .block-container {
    padding: 24px 28px !important;
    max-width: 100% !important;
}

/* ── CARDS MÉTRICA ── */
.metric-card {
    background: white;
    border-radius: 10px;
    padding: 20px 22px;
    border: 1px solid #E5E9F0;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
    transition: box-shadow 0.2s;
}
.metric-card:hover { box-shadow: 0 4px 16px rgba(0,0,0,0.10); }
.metric-label {
    font-size: 12px; font-weight: 600; color: #8896A7;
    text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 8px;
}
.metric-value {
    font-size: 32px; font-weight: 800; color: #1A3A5C; line-height: 1;
}
.metric-delta {
    font-size: 12px; font-weight: 600; margin-top: 6px;
    display: flex; align-items: center; gap: 4px;
}
.delta-up { color: #2ECC71; }
.delta-down { color: #E74C3C; }

/* ── CARDS PAINEL ── */
.panel-card {
    background: white;
    border-radius: 10px;
    padding: 24px;
    border: 1px solid #E5E9F0;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
    height: 100%;
}
.panel-title {
    font-size: 15px; font-weight: 700; color: #1A3A5C;
    margin-bottom: 20px;
}

/* ── TABELA ── */
[data-testid="stDataFrame"] {
    border: 1px solid #E5E9F0 !important;
    border-radius: 10px !important;
    overflow: hidden !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.04) !important;
    background: white !important;
}
[data-testid="stDataFrame"] table { background: white !important; }
[data-testid="stDataFrame"] thead th {
    background: #F4F6F9 !important;
    color: #8896A7 !important;
    font-size: 11px !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.05em !important;
}
[data-testid="stDataFrame"] tbody tr:hover td {
    background: #F8FAFC !important;
}

/* ── INPUTS ── */
.stTextInput input, .stNumberInput input, .stTextArea textarea,
.stSelectbox > div > div, [data-testid="stSelectbox"] > div > div {
    background: white !important;
    border: 1px solid #D1D9E6 !important;
    border-radius: 8px !important;
    color: #1A3A5C !important;
    font-size: 14px !important;
}
.stTextInput input:focus, .stNumberInput input:focus {
    border-color: #1A7FC1 !important;
    box-shadow: 0 0 0 3px rgba(26,127,193,0.15) !important;
}

/* ── BOTÕES ── */
.stButton > button {
    background: #1A7FC1 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: 14px !important;
    padding: 10px 20px !important;
    transition: all 0.15s !important;
}
.stButton > button:hover {
    background: #1565A0 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 12px rgba(26,127,193,0.3) !important;
}
button[kind="secondary"] {
    background: white !important;
    color: #E74C3C !important;
    border: 1px solid #E74C3C !important;
}

/* ── DOWNLOAD BUTTON ── */
.stDownloadButton > button {
    background: white !important;
    color: #2ECC71 !important;
    border: 1px solid #2ECC71 !important;
    border-radius: 8px !important;
}

/* ── FORMS ── */
[data-testid="stForm"] {
    background: white !important;
    border: 1px solid #E5E9F0 !important;
    border-radius: 12px !important;
    padding: 24px !important;
}

/* ── EXPANDER padrão (branco) ── */
.streamlit-expanderHeader {
    background: white !important;
    border: 1px solid #E5E9F0 !important;
    border-radius: 8px !important;
    color: #1A3A5C !important;
    font-weight: 600 !important;
}
.streamlit-expanderContent {
    background: #FAFBFC !important;
    border: 1px solid #E5E9F0 !important;
}

/* ── EXPANDER DARK (IA) ── */
.expander-dark .streamlit-expanderHeader,
[data-testid="stExpander"].dark-expander > div:first-child {
    background: #1A3A5C !important;
    border: 1px solid #2A5080 !important;
    border-radius: 8px !important;
    color: white !important;
}
.expander-dark .streamlit-expanderContent,
[data-testid="stExpander"].dark-expander > div:last-child {
    background: #0F2540 !important;
    border: 1px solid #2A5080 !important;
}

/* Período dentro da aba — inputs brancos menores */
.periodo-inline .stSelectbox > div > div {
    background: white !important;
    border: 1px solid #D1D9E6 !important;
    font-size: 13px !important;
}


/* ── ALERTS ── */
.stAlert { border-radius: 8px !important; }
[data-testid="stSuccess"] { background: #F0FFF4 !important; border-color: #2ECC71 !important; color: #1A6B3A !important; }
[data-testid="stError"] { background: #FFF5F5 !important; border-color: #E74C3C !important; color: #6B1A1A !important; }
[data-testid="stInfo"] { background: #EFF8FF !important; border-color: #1A7FC1 !important; color: #1A3A5C !important; }
[data-testid="stWarning"] { background: #FFFBEB !important; border-color: #F39C12 !important; color: #6B4A00 !important; }

/* ── PÁGINA ── */
h1 { color: #1A3A5C !important; font-size: 22px !important; font-weight: 800 !important; margin-bottom: 4px !important; }
h2, h3 { color: #1A3A5C !important; }
p, .stCaption { color: #5A6A7A !important; }

/* Esconder hamburger */
[data-testid="collapsedControl"] { display: none !important; }

/* Divider */
hr { border-color: #E5E9F0 !important; }

/* Status badges coloridos */
.badge-aberto   { background:#FEF3F2;color:#E74C3C;border:1px solid #FECACA;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700 }
.badge-adiantado{ background:#FFF7ED;color:#EA8C00;border:1px solid #FED7AA;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700 }
.badge-pendente { background:#F5F3FF;color:#7C3AED;border:1px solid #DDD6FE;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700 }
.badge-fechado  { background:#F0FFF4;color:#2ECC71;border:1px solid #BBF7D0;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700 }
</style>
""", unsafe_allow_html=True)

# ── SUPABASE ───────────────────────────────────────────────────────
SUPABASE_URL = "https://lmcefcmjatnixrsggyvz.supabase.co"
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
ANTHROPIC_KEY = st.secrets.get("ANTHROPIC_KEY", "")

HDR = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
       "Content-Type": "application/json", "Prefer": "return=representation"}

def sb_get(t, p=""):
    try:
        r = requests.get(f"{SUPABASE_URL}/rest/v1/{t}?{p}", headers=HDR, timeout=10)
        return r.json() if r.ok else []
    except: return []

def sb_post(t, d, upsert=False):
    h = {**HDR, "Prefer": "resolution=merge-duplicates,return=representation" if upsert else "return=representation"}
    try:
        r = requests.post(f"{SUPABASE_URL}/rest/v1/{t}", json=d, headers=h, timeout=10)
        return r.ok
    except: return False

def sb_patch(t, f, d):
    try:
        r = requests.patch(f"{SUPABASE_URL}/rest/v1/{t}?{f}", json=d, headers=HDR, timeout=10)
        return r.ok
    except: return False

def sb_delete(t, f):
    try:
        r = requests.delete(f"{SUPABASE_URL}/rest/v1/{t}?{f}", headers={**HDR, "Prefer": ""}, timeout=10)
        return r.ok
    except: return False

# ── LEITURA IA ─────────────────────────────────────────────────────
def ler_contrato_ia(img_bytes, media_type="image/jpeg"):
    if not ANTHROPIC_KEY:
        return None, "⚠️ Chave ANTHROPIC_KEY não configurada nos Secrets do Streamlit."
    b64 = base64.b64encode(img_bytes).decode("utf-8")
    prompt = """Analise este contrato de transporte e retorne APENAS um JSON puro, sem markdown, sem explicação:
{"cliente":"AUTOPORT","motorista":"NOME EM MAIUSCULAS","placa":"ABC1234","frota":"","contrato":"numero","data":"DD/MM/AAAA","fat_bruto":0.0,"chapa":0.0,"destino":"Cidade/UF","qtd_veiculos":0,"dt_pagamento":"DD/MM/AAAA"}
Retorne SOMENTE o JSON."""
    is_pdf = media_type == "application/pdf"
    hdrs = {"x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01", "content-type": "application/json"}
    if is_pdf: hdrs["anthropic-beta"] = "pdfs-2024-09-25"
    bloco = {"type": "document" if is_pdf else "image",
             "source": {"type": "base64", "media_type": media_type, "data": b64}}
    try:
        r = requests.post("https://api.anthropic.com/v1/messages", headers=hdrs,
            json={"model": "claude-sonnet-4-20250514", "max_tokens": 1024,
                  "messages": [{"role": "user", "content": [bloco, {"type": "text", "text": prompt}]}]},
            timeout=60)
        if not r.ok:
            resp = r.json() if "application/json" in r.headers.get("content-type","") else {}
            return None, f"Erro API ({r.status_code}): {resp.get('error',{}).get('message', r.text[:200])}"
        data = r.json()
        texto = data.get("content", [{}])[0].get("text", "")
        clean = texto.replace("```json","").replace("```","").strip()
        s = clean.find("{"); e = clean.rfind("}") + 1
        if s >= 0 and e > s: return json.loads(clean[s:e]), None
        return None, f"JSON não encontrado: {clean[:200]}"
    except requests.exceptions.Timeout:
        return None, "Timeout: arquivo pode ser grande demais ou rede lenta."
    except Exception as ex:
        return None, str(ex)

# ── CONSTANTES ─────────────────────────────────────────────────────
MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
META = 127000
SEM = ["FLAVIO","MARIO","ORANGE CARVALHO","WEMERSON CARLOS"]
MOTORISTAS = ["ALEX","DAIVDSON","ELEXSANDRO","GEOVANE","FLAVIO","HEBERT","JOSÉ EDUARDO","LUIZ OTAVIO","MARIO","REINALDO ADRIANO","ROBINSON TAVARES","WAGNER","WEMERSON CARLOS","ORANGE CARVALHO"]
CLIENTES = ["SADA","AUTOPORT","DACUNHA","BRAZUL","VIX","TRANSAUTO","TRANSZERO","OUTRO"]
STATUS = ["ABERTO","ADIANTADO","PENDENTE","FECHADO"]
STATUS_P = ["QUALIFICADO","EM ANÁLISE","APROVADO","PAGO","NÃO APROVADO"]
PERFIS = {
    "🏢 Escritório": {"senha": "roesel2026", "perm": "total"},
    "👩‍💼 Claudiane": {"senha": "claudiane123", "perm": "view"},
    "👥 Equipe": {"senha": "equipe2026", "perm": "equipe"}
}
STATUS_COR = {"ABERTO":"#E74C3C","ADIANTADO":"#EA8C00","PENDENTE":"#7C3AED","FECHADO":"#2ECC71"}

# Cores do gráfico: tema claro/profissional
PLOT = dict(
    paper_bgcolor="white", plot_bgcolor="white",
    font=dict(color="#5A6A7A", family="Inter, sans-serif", size=12),
    xaxis=dict(gridcolor="#F0F2F5", color="#8896A7", showline=False, tickfont=dict(size=11)),
    yaxis=dict(gridcolor="#F0F2F5", color="#8896A7", showline=False, tickfont=dict(size=11)),
    margin=dict(l=0, r=20, t=20, b=0)
)

def R(v):
    try: return f"R$ {float(v or 0):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    except: return "R$ 0,00"

def R_short(v):
    """Formata valor de forma resumida: R$ 127k, R$ 1,2M"""
    try:
        f = float(v or 0)
        if f >= 1_000_000: return f"R$ {f/1_000_000:.1f}M".replace(".",",")
        if f >= 1_000: return f"R$ {f/1_000:.1f}k".replace(".",",")
        return R(f)
    except: return "R$ 0"

def com(m, f):
    f = float(f or 0)
    return (0, f*0.10) if m in SEM else (f*0.05, f*0.05)

def fd(d):
    try: return datetime.strptime(str(d)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except: return ""

def badge_html(status):
    cls = {"ABERTO":"badge-aberto","ADIANTADO":"badge-adiantado","PENDENTE":"badge-pendente","FECHADO":"badge-fechado"}.get(status,"badge-aberto")
    return f'<span class="{cls}">{status}</span>'

def gerar_excel(df, titulo="Relatório"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = titulo[:31]
    thin = Side(style="thin", color="E0E0E0")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Carlos Roesel Transportes — {titulo}"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor="1A3A5C")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="888888", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    colunas = ["MOTORISTA","CLIENTE","CONTRATO","DATA","FAT. BRUTO","CHAPA","ADIANTAMENTO","FOLHA","DESTINO","STATUS"]
    for i, col in enumerate(colunas, 1):
        c = ws.cell(row=3, column=i, value=col)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor="1A7FC1")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda
    ws.row_dimensions[3].height = 20
    df2 = df.copy()
    if "data" in df2.columns:
        df2["data"] = pd.to_datetime(df2["data"], errors="coerce").dt.strftime("%d/%m/%Y")
    for ri, (_, row) in enumerate(df2.iterrows(), 4):
        fill_cor = "FAFBFC" if ri % 2 == 0 else "FFFFFF"
        a_val, f_val = com(row.get("motorista",""), row.get("fat_bruto", 0))
        linha = [row.get("motorista",""), row.get("cliente",""), row.get("contrato",""),
                 row.get("data",""), float(row.get("fat_bruto",0)), float(row.get("chapa",0)),
                 a_val if row.get("motorista","") not in SEM else 0, f_val,
                 row.get("destino",""), row.get("status","")]
        for ci, val in enumerate(linha, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=fill_cor)
            c.border = borda
            if ci in (5,6,7,8): c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal="right")
            elif ci == 4: c.alignment = Alignment(horizontal="center")
            else: c.alignment = Alignment(horizontal="left")
    for ci, w in enumerate([20,14,16,12,16,14,18,14,20,12], 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ══════════════════════════════════════════════════════════════════
# LOGIN
# ══════════════════════════════════════════════════════════════════
if "ok" not in st.session_state:
    st.session_state.ok = False

if not st.session_state.ok:
    # Centraliza o login
    col_l, col_m, col_r = st.columns([1, 1, 1])
    with col_m:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("""
        <div style='background:white;border-radius:16px;padding:48px 40px;border:1px solid #E5E9F0;box-shadow:0 8px 40px rgba(26,58,92,0.12);text-align:center'>
            <div style='font-size:52px;margin-bottom:12px'>🚛</div>
            <div style='font-size:20px;font-weight:800;color:#1A3A5C;letter-spacing:0.08em;text-transform:uppercase'>Carlos Roesel</div>
            <div style='font-size:13px;font-weight:600;color:#1A7FC1;letter-spacing:0.12em;text-transform:uppercase;margin-bottom:4px'>TRANSPORTES</div>
            <div style='font-size:11px;color:#B8CDE0;margin-bottom:32px'>Sistema de Gestão de Contratos</div>
            <hr style='border-color:#F0F2F5;margin-bottom:28px'>
        </div>
        """, unsafe_allow_html=True)
        p = st.selectbox("Perfil", list(PERFIS.keys()))
        s = st.text_input("Senha", type="password", placeholder="Digite sua senha")
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Acessar o sistema →", use_container_width=True):
            if s == PERFIS[p]["senha"]:
                st.session_state.ok = True
                st.session_state.perfil = p
                st.session_state.perm = PERFIS[p]["perm"]
                st.rerun()
            else:
                st.error("Senha incorreta.")
    st.stop()

perm = st.session_state.perm

# ── DADOS ──────────────────────────────────────────────────────────
@st.cache_data(ttl=15)
def get_df():
    d = sb_get("contratos", "order=data.desc")
    if not d or not isinstance(d, list): return pd.DataFrame()
    df = pd.DataFrame(d)
    if df.empty: return df
    df["data"] = pd.to_datetime(df["data"], errors="coerce")
    for c in ["fat_bruto","chapa"]:
        df[c] = pd.to_numeric(df.get(c, 0), errors="coerce").fillna(0)
    df["adiant"] = df.apply(lambda r: com(r["motorista"], r["fat_bruto"])[0], axis=1)
    df["folha"]  = df.apply(lambda r: com(r["motorista"], r["fat_bruto"])[1], axis=1)
    return df

@st.cache_data(ttl=15)
def get_premios():
    d = sb_get("premios")
    return {p["motorista"]: p for p in d} if d and isinstance(d, list) else {}

df_all   = get_df()
prem_map = get_premios()

# ══════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════
with st.sidebar:
    # Logo / Marca
    st.markdown("""
    <div style='padding:20px 16px 16px;border-bottom:1px solid rgba(255,255,255,0.1);margin-bottom:8px'>
        <div style='font-size:26px;font-weight:900;color:white;letter-spacing:0.08em'>🚛 RT</div>
        <div style='font-size:10px;color:#7AADCC;letter-spacing:0.15em;text-transform:uppercase;margin-top:2px'>Roesel Transportes</div>
    </div>
    """, unsafe_allow_html=True)

    # Menu
    st.markdown("<div style='padding:8px 12px 4px;font-size:10px;font-weight:700;color:#5A8BAD;text-transform:uppercase;letter-spacing:0.1em'>Menu</div>", unsafe_allow_html=True)

    ABAS = {
        "📊  Visão Geral": "dashboard",
        "➕  Novo Contrato": "novo",
        "📋  Contratos": "contratos",
        "👤  Por Motorista": "motorista",
        "💳  Comissões": "comissoes",
        "🏆  Prêmios": "premios",
    }
    if "aba" not in st.session_state:
        st.session_state.aba = "dashboard"

    for label, key in ABAS.items():
        ativo = st.session_state.aba == key
        bg = "rgba(255,255,255,0.15)" if ativo else "transparent"
        cor = "white" if ativo else "#B8CDE0"
        borda_l = "3px solid #4DB8FF" if ativo else "3px solid transparent"
        if st.button(label, key=f"menu_{key}", use_container_width=True):
            st.session_state.aba = key
            st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)
    st.divider()

    # Usuário logado
    perfil_nome = st.session_state.perfil
    st.markdown(f"""
    <div style='padding:12px;background:rgba(255,255,255,0.07);border-radius:8px;margin:4px 0'>
        <div style='font-size:11px;color:#7AADCC;margin-bottom:2px'>Logado como</div>
        <div style='font-size:13px;color:white;font-weight:600'>{perfil_nome}</div>
    </div>
    """, unsafe_allow_html=True)
    if st.button("🚪 Sair", use_container_width=True):
        st.session_state.ok = False
        st.rerun()

# ── SELETOR DE PERÍODO INLINE ─────────────────────────────────────
def sel_periodo(key_suffix=""):
    """Renderiza seletor de Ano/Mês inline dentro da aba."""
    # Anos: combina anos do banco + anos fixos até 2030, sem duplicatas
    anos_banco = set(df_all["data"].dt.year.dropna().unique().astype(int).tolist()) if not df_all.empty else set()
    anos_fixos = set(range(2024, 2031))
    anos_list = [0] + sorted(anos_banco | anos_fixos, reverse=True)

    c1, c2, c3 = st.columns([1, 1, 4])
    with c1:
        ano = st.selectbox("Ano", anos_list,
            index=anos_list.index(datetime.now().year) if datetime.now().year in anos_list else 1,
            format_func=lambda x: "Todos os anos" if x == 0 else str(x),
            key=f"ano_{key_suffix}")
    with c2:
        mes = st.selectbox("Mês", [0] + list(range(1, 13)),
            index=datetime.now().month,
            format_func=lambda x: "Todos" if x == 0 else MESES[x - 1],
            key=f"mes_{key_suffix}")
    periodo = f"{MESES[mes-1]}/{ano}" if mes and ano else \
              (MESES[mes-1] if mes else (str(ano) if ano else "Todos os períodos"))
    dfp = df_all.copy()
    if not df_all.empty:
        if ano: dfp = dfp[dfp["data"].dt.year == ano]
        if mes: dfp = dfp[dfp["data"].dt.month == mes]
    return dfp, periodo

aba = st.session_state.aba

# ══════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════
if aba == "dashboard":
    # Cabeçalho + seletor período
    col_h1, col_h2 = st.columns([3, 1])
    with col_h1:
        st.markdown("<h1 style='margin:0 0 4px'>Visão Geral</h1>", unsafe_allow_html=True)
    with col_h2:
        st.markdown(f"<div style='text-align:right;padding:6px 0;font-size:12px;color:#8896A7'>🕐 {datetime.now().strftime('%d/%m %H:%M')}</div>", unsafe_allow_html=True)

    df, periodo = sel_periodo("dash")
    st.markdown(f"<p style='color:#8896A7;font-size:13px;margin:-4px 0 20px'>📅 {periodo} &nbsp;·&nbsp; {len(df)} contratos</p>", unsafe_allow_html=True)

    if df.empty:
        st.info(f"Nenhum dado encontrado para {periodo}.")
    else:
        fat = df["fat_bruto"].sum()
        pend = len(df[df["status"].isin(["ABERTO","PENDENTE"])])
        meta_pct = min(fat / META * 100, 100)

        # ── MÉTRICAS TOP ──
        c1, c2, c3, c4, c5 = st.columns(5)
        metricas = [
            (c1, "💰 Faturamento Total", R_short(fat), f"{meta_pct:.0f}% da meta", meta_pct >= 80, "#1A7FC1"),
            (c2, "📄 Total de Contratos", str(len(df)), f"{pend} pendentes", pend == 0, "#8B5CF6"),
            (c3, "⏩ Adiantamentos", R_short(df["adiant"].sum()), "pago aos motoristas", True, "#F59E0B"),
            (c4, "📋 Folha Total", R_short(df["folha"].sum()), "comissão motoristas", True, "#10B981"),
            (c5, "🔧 Total Chapas", R_short(df["chapa"].sum()), "custo chapas", True, "#EC4899"),
        ]
        for col, label, valor, sub, positivo, cor in metricas:
            with col:
                st.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-label'>{label}</div>
                    <div class='metric-value' style='color:{cor}'>{valor}</div>
                    <div class='metric-delta {"delta-up" if positivo else "delta-down"}'>
                        {"▲" if positivo else "▼"} {sub}
                    </div>
                </div>
                """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── GRÁFICO DE BARRAS + PIZZA ──
        col_g1, col_g2 = st.columns([3, 1.8])

        with col_g1:
            st.markdown("""<div class='panel-card'>
                <div class='panel-title'>📈 Faturamento por Motorista</div>""",
                unsafe_allow_html=True)
            fm = df.groupby("motorista")["fat_bruto"].sum().sort_values().reset_index()
            fm["cor"] = fm["fat_bruto"].apply(lambda x: "#2ECC71" if x >= META else "#1A7FC1")
            fig = go.Figure(go.Bar(
                x=fm["fat_bruto"], y=fm["motorista"], orientation="h",
                marker_color=fm["cor"],
                text=[R_short(v) for v in fm["fat_bruto"]],
                textposition="outside",
                textfont=dict(color="#5A6A7A", size=11)
            ))
            fig.add_vline(x=META, line_dash="dash", line_color="#F59E0B", line_width=2,
                annotation_text="Meta", annotation_font_color="#F59E0B", annotation_font_size=11)
            fig.update_layout(height=380, showlegend=False, **PLOT)
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
            st.markdown("</div>", unsafe_allow_html=True)

        with col_g2:
            st.markdown("""<div class='panel-card'>
                <div class='panel-title'>🥧 Status dos Contratos</div>""",
                unsafe_allow_html=True)
            sc = df["status"].value_counts().reset_index()
            CORES_PIE = {"ABERTO":"#E74C3C","ADIANTADO":"#F59E0B","PENDENTE":"#8B5CF6","FECHADO":"#2ECC71"}
            fig2 = px.pie(sc, values="count", names="status",
                color="status", color_discrete_map=CORES_PIE, hole=0.55)
            fig2.update_layout(height=380,
                paper_bgcolor="white", plot_bgcolor="white",
                font=dict(color="#5A6A7A", size=12),
                legend=dict(orientation="v", yanchor="middle", y=0.5, font=dict(size=12)),
                margin=dict(l=0, r=10, t=10, b=10))
            fig2.update_traces(textfont_color="white", textfont_size=12)
            st.plotly_chart(fig2, use_container_width=True, config={"displayModeBar": False})
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── PAINÉIS INFERIORES ──
        col_p1, col_p2, col_p3 = st.columns([1.4, 1.3, 1.3])

        with col_p1:
            st.markdown("""<div class='panel-card'>
                <div class='panel-title'>🏆 Top Motoristas</div>""",
                unsafe_allow_html=True)
            rk = df.groupby("motorista")["fat_bruto"].sum().sort_values(ascending=False).head(5).reset_index()
            st.markdown("<table style='width:100%;border-collapse:collapse'>", unsafe_allow_html=True)
            st.markdown("<tr><th style='text-align:left;padding:6px;font-size:11px;color:#8896A7;font-weight:700;border-bottom:2px solid #F0F2F5'>MOTORISTA</th><th style='text-align:right;padding:6px;font-size:11px;color:#8896A7;font-weight:700;border-bottom:2px solid #F0F2F5'>FATURAMENTO</th><th style='text-align:center;padding:6px;font-size:11px;color:#8896A7;font-weight:700;border-bottom:2px solid #F0F2F5'>META</th></tr>",
                unsafe_allow_html=True)
            medalhas = ["🥇","🥈","🥉","4°","5°"]
            for i, row in rk.iterrows():
                el = row["fat_bruto"] >= META
                cor_val = "#2ECC71" if el else "#1A3A5C"
                st.markdown(f"""
                <tr style='border-bottom:1px solid #F4F6F9'>
                    <td style='padding:10px 6px;font-size:13px;color:#1A3A5C;font-weight:600'>
                        {medalhas[i]} {row['motorista']}
                    </td>
                    <td style='padding:10px 6px;text-align:right;font-size:13px;font-weight:700;color:{cor_val}'>
                        {R_short(row['fat_bruto'])}
                    </td>
                    <td style='padding:10px 6px;text-align:center'>
                        {"✅" if el else f"<span style='font-size:11px;color:#EA8C00'>{min(row['fat_bruto']/META*100,100):.0f}%</span>"}
                    </td>
                </tr>""", unsafe_allow_html=True)
            st.markdown("</table></div>", unsafe_allow_html=True)

        with col_p2:
            st.markdown("""<div class='panel-card'>
                <div class='panel-title'>📦 Contratos por Cliente</div>""",
                unsafe_allow_html=True)
            cli_fat = df.groupby("cliente")["fat_bruto"].sum().sort_values(ascending=False).head(5)
            total_cli = cli_fat.sum()
            for cli, val in cli_fat.items():
                pct = val / total_cli * 100 if total_cli > 0 else 0
                st.markdown(f"""
                <div style='margin-bottom:14px'>
                    <div style='display:flex;justify-content:space-between;margin-bottom:5px'>
                        <span style='font-size:13px;font-weight:600;color:#1A3A5C'>{cli}</span>
                        <span style='font-size:12px;color:#5A6A7A;font-weight:500'>{R_short(val)} <span style='color:#8896A7'>({pct:.0f}%)</span></span>
                    </div>
                    <div style='background:#F4F6F9;border-radius:4px;height:6px'>
                        <div style='background:#1A7FC1;width:{pct}%;height:6px;border-radius:4px'></div>
                    </div>
                </div>""", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

        with col_p3:
            st.markdown("""<div class='panel-card'>
                <div class='panel-title'>📊 Resumo Financeiro</div>""",
                unsafe_allow_html=True)
            total_adiant = df["adiant"].sum()
            total_folha  = df["folha"].sum()
            total_chapa  = df["chapa"].sum()
            total_sai    = total_adiant + total_folha + total_chapa
            itens = [
                ("💰 Faturamento Bruto", fat, "#1A7FC1"),
                ("⏩ Adiantamentos", total_adiant, "#F59E0B"),
                ("📋 Folha de Pagto.", total_folha, "#8B5CF6"),
                ("🔧 Chapas", total_chapa, "#EC4899"),
                ("📤 Total de Saídas", total_sai, "#E74C3C"),
                ("✅ Saldo Líquido", fat - total_sai, "#2ECC71"),
            ]
            for label, val, cor in itens:
                negrito = "800" if label in ("✅ Saldo Líquido", "💰 Faturamento Bruto") else "500"
                st.markdown(f"""
                <div style='display:flex;justify-content:space-between;padding:9px 0;border-bottom:1px solid #F4F6F9'>
                    <span style='font-size:13px;color:#5A6A7A;font-weight:{negrito}'>{label}</span>
                    <span style='font-size:13px;font-weight:700;color:{cor}'>{R_short(val)}</span>
                </div>""", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════
# NOVO CONTRATO
# ══════════════════════════════════════════════════════════════════
elif aba == "novo":
    if perm not in ["total","equipe"]:
        st.warning("Sem permissão para cadastrar contratos.")
        st.stop()

    st.markdown("<h1>➕ Novo Contrato</h1><p style='color:#8896A7;margin-top:-8px'>Cadastre um novo contrato de transporte</p>", unsafe_allow_html=True)

    # CSS escuro só para o expander da IA
    st.markdown("""
    <style>
    [data-testid="stExpander"]:first-of-type > div:first-child {
        background: #1A3A5C !important;
        border: 1px solid #2A5A8C !important;
        border-radius: 10px !important;
    }
    [data-testid="stExpander"]:first-of-type > div:first-child p,
    [data-testid="stExpander"]:first-of-type > div:first-child span,
    [data-testid="stExpander"]:first-of-type > div:first-child svg {
        color: white !important;
        fill: white !important;
    }
    [data-testid="stExpander"]:first-of-type > div:last-child {
        background: #0F2540 !important;
        border: 1px solid #2A5A8C !important;
        border-top: none !important;
        border-radius: 0 0 10px 10px !important;
        padding: 20px !important;
    }
    [data-testid="stExpander"]:first-of-type > div:last-child p,
    [data-testid="stExpander"]:first-of-type > div:last-child label,
    [data-testid="stExpander"]:first-of-type > div:last-child span:not([class*="badge"]) {
        color: #B8CDE0 !important;
    }
    [data-testid="stExpander"]:first-of-type input,
    [data-testid="stExpander"]:first-of-type textarea,
    [data-testid="stExpander"]:first-of-type [data-baseweb="select"] > div {
        background: #1E4976 !important;
        border-color: #2A5A8C !important;
        color: white !important;
    }
    [data-testid="stExpander"]:first-of-type .stButton > button {
        background: #1A7FC1 !important;
        color: white !important;
    }
    [data-testid="stExpander"]:first-of-type [data-testid="stAlert"] {
        background: rgba(255,255,255,0.08) !important;
        border-color: rgba(255,255,255,0.2) !important;
        color: #B8CDE0 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    with st.expander("📷 Importar via IA (Foto / PDF)", expanded=False):
        if not ANTHROPIC_KEY:
            st.warning("⚠️ Configure `ANTHROPIC_KEY` nos Secrets do Streamlit para usar a leitura automática.")
        else:
            st.markdown("<p style='color:#7AADCC;font-size:13px'>Envie uma foto ou PDF do contrato — a IA preenche os campos automaticamente.</p>", unsafe_allow_html=True)
        upl = st.file_uploader("Arquivo do contrato", type=["jpg","jpeg","png","webp","pdf"])
        if upl:
            col_a, col_b = st.columns([2,1])
            with col_a:
                if upl.type != "application/pdf":
                    st.image(upl, use_container_width=True)
                else:
                    st.info(f"📄 **{upl.name}** — PDF pronto para análise")
            with col_b:
                st.markdown("<br><br>", unsafe_allow_html=True)
                if st.button("🤖 Analisar com IA", use_container_width=True, disabled=not ANTHROPIC_KEY):
                    arquivo = upl.read()
                    if len(arquivo) > 5 * 1024 * 1024:
                        st.error("Arquivo muito grande (máx. 5 MB).")
                    else:
                        with st.spinner("Analisando…"):
                            dados, erro = ler_contrato_ia(arquivo, upl.type)
                        if erro:
                            st.error(f"Erro: {erro}")
                        elif dados:
                            st.session_state["ia"] = dados
                            st.success("✅ Dados extraídos com sucesso!")
                            st.rerun()

    ia = st.session_state.get("ia", {})
    if ia:
        st.success("📋 Campos preenchidos pela IA — confira e salve.")

    with st.form("fnovo", clear_on_submit=True):
        st.markdown("**Dados do Contrato**")
        c1,c2,c3 = st.columns(3)
        mi = MOTORISTAS.index(ia.get("motorista","")) if ia.get("motorista","") in MOTORISTAS else 0
        ci = CLIENTES.index(ia.get("cliente","")) if ia.get("cliente","") in CLIENTES else 0
        mot   = c1.selectbox("Motorista *", MOTORISTAS, index=mi)
        cli   = c2.selectbox("Cliente *", CLIENTES, index=ci)
        placa = c3.text_input("Placa *", ia.get("placa","")).upper()
        c4,c5,c6 = st.columns(3)
        cont  = c4.text_input("Nº Contrato *", ia.get("contrato",""))
        frota = c5.text_input("Frota", ia.get("frota",""))
        dia   = datetime.now()
        if ia.get("data"):
            try: dia = datetime.strptime(ia["data"], "%d/%m/%Y")
            except: pass
        data_v = c6.date_input("Data *", dia)
        st.markdown("**Valores**")
        c7,c8,c9 = st.columns(3)
        fat_v   = c7.number_input("Fat. Bruto (R$) *", float(ia.get("fat_bruto",0)), step=100.0, format="%.2f")
        chapa_v = c8.number_input("Chapa (R$)", float(ia.get("chapa",0)), step=50.0, format="%.2f")
        qtd     = c9.number_input("Qtd Veículos", int(ia.get("qtd_veiculos",0)), step=1)
        st.markdown("**Informações Adicionais**")
        c10,c11 = st.columns(2)
        dest   = c10.text_input("Destino", ia.get("destino","")).upper()
        sts    = c11.selectbox("Status", STATUS)
        c12,c13 = st.columns(2)
        dt_pag = c12.date_input("Dt. Pagamento", value=None)
        adpago = c13.checkbox("Adiantamento Pago?")
        obs    = st.text_area("Observação", "", height=70)
        if fat_v > 0:
            a, f = com(mot, fat_v)
            st.markdown(f"""
            <div style='background:#EFF8FF;border:1px solid #BFDBFE;border-radius:8px;padding:12px 16px;font-size:13px;color:#1A3A5C;margin:8px 0'>
                💳 {"Sem adiantamento" if mot in SEM else f"Adiantamento: <b style='color:#F59E0B'>{R(a)}</b>"} &nbsp;·&nbsp; Folha: <b style='color:#8B5CF6'>{R(f)}</b>
            </div>""", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.form_submit_button("✅ Salvar Contrato", use_container_width=True):
            if not cont or not fat_v:
                st.error("Preencha: Nº Contrato e Faturamento Bruto.")
            else:
                novo = {"id": str(uuid.uuid4()), "motorista": mot, "cliente": cli, "placa": placa,
                        "frota": frota, "contrato": cont, "data": str(data_v), "fat_bruto": fat_v,
                        "chapa": chapa_v, "destino": dest, "qtd_veiculos": int(qtd),
                        "adiantamento_pago": adpago,
                        "dt_pagamento": str(dt_pag) if dt_pag else None,
                        "status": sts, "obs": obs}
                if sb_post("contratos", novo):
                    st.success(f"✅ Contrato {cont} salvo!")
                    st.cache_data.clear()
                    st.session_state.pop("ia", None)
                    st.rerun()
                else:
                    st.error("❌ Erro ao salvar. Verifique a conexão com o Supabase.")

# ══════════════════════════════════════════════════════════════════
# CONTRATOS
# ══════════════════════════════════════════════════════════════════
elif aba == "contratos":
    st.markdown("<h1>📋 Contratos</h1>", unsafe_allow_html=True)
    df, periodo = sel_periodo("cont")
    st.markdown(f"<p style='color:#8896A7;margin-top:-8px'>📅 {periodo}</p>", unsafe_allow_html=True)

    c1,c2,c3 = st.columns(3)
    busca = c1.text_input("🔍 Buscar", "", placeholder="Motorista, nº contrato, destino…")
    fs = c2.selectbox("Filtrar por Status", ["Todos"] + STATUS)
    fc = c3.selectbox("Filtrar por Cliente", ["Todos"] + CLIENTES)

    dv = df.copy()
    if not dv.empty:
        if busca:
            b = busca.upper()
            dv = dv[dv.apply(lambda r:
                b in str(r.get("motorista","")).upper() or
                b in str(r.get("contrato","")).upper() or
                b in str(r.get("destino","")).upper(), axis=1)]
        if fs != "Todos": dv = dv[dv["status"] == fs]
        if fc != "Todos": dv = dv[dv.get("cliente","") == fc]

    # Sumário em cards pequenos
    col_s1,col_s2,col_s3,col_s4 = st.columns(4)
    for col, label, val, cor in [
        (col_s1, "Contratos", str(len(dv)), "#1A7FC1"),
        (col_s2, "Faturamento", R_short(dv["fat_bruto"].sum() if not dv.empty else 0), "#2ECC71"),
        (col_s3, "Adiantamentos", R_short(dv["adiant"].sum() if not dv.empty else 0), "#F59E0B"),
        (col_s4, "Folha", R_short(dv["folha"].sum() if not dv.empty else 0), "#8B5CF6"),
    ]:
        with col:
            st.markdown(f"""<div class='metric-card' style='padding:14px 18px'>
                <div class='metric-label'>{label}</div>
                <div style='font-size:22px;font-weight:800;color:{cor}'>{val}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    if not dv.empty:
        cols = [c for c in ["motorista","cliente","contrato","data","fat_bruto","chapa","destino","status","adiantamento_pago"] if c in dv.columns]
        ds = dv[cols].copy()
        ds["data"] = ds["data"].dt.strftime("%d/%m/%Y")
        ds["fat_bruto"] = ds["fat_bruto"].apply(R)
        ds["chapa"] = ds["chapa"].apply(R)
        ds["adiantamento_pago"] = ds["adiantamento_pago"].apply(lambda x: "✅" if x else "❌")
        ds.columns = ["MOTORISTA","CLIENTE","CONTRATO","DATA","FAT. BRUTO","CHAPA","DESTINO","STATUS","ADIANT."][:len(cols)]
        st.dataframe(ds, use_container_width=True, hide_index=True,
            column_config={
                "MOTORISTA": st.column_config.TextColumn(width="medium"),
                "FAT. BRUTO": st.column_config.TextColumn(width="medium"),
                "STATUS": st.column_config.TextColumn(width="small"),
                "ADIANT.": st.column_config.TextColumn(width="small"),
            })

        # Exportações
        col_e1, col_e2 = st.columns(2)
        with col_e1:
            csv = dv.to_csv(index=False, sep=";", encoding="utf-8-sig")
            st.download_button("📥 Exportar CSV", csv,
                f"contratos_{periodo.replace('/','_')}.csv", "text/csv",
                use_container_width=True)
        with col_e2:
            try:
                excel_buf = gerar_excel(dv, f"Contratos {periodo}")
                st.download_button("📊 Exportar Excel (.xlsx)", excel_buf,
                    f"contratos_{periodo.replace('/','_')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            except Exception as ex:
                st.caption(f"Excel indisponível: {ex}")

        if perm == "total":
            st.markdown("---")
            st.markdown("#### ✏️ Editar / Excluir")
            labels = [f"{r['motorista']} — {r['contrato']} — {fd(str(r['data'])[:10])}" for _, r in dv.iterrows()]
            sel = st.selectbox("Selecione o contrato", labels)
            row = dv.iloc[labels.index(sel)]
            with st.expander("✏️ Editar contrato"):
                with st.form("fedit"):
                    ec1,ec2 = st.columns(2)
                    es = ec1.selectbox("Status", STATUS,
                        index=STATUS.index(row.get("status","ABERTO")) if row.get("status") in STATUS else 0)
                    ea = ec2.checkbox("Adiantamento Pago?", value=bool(row.get("adiantamento_pago")))
                    eo = st.text_area("Obs", value=row.get("obs","") or "")
                    if st.form_submit_button("💾 Salvar"):
                        if sb_patch("contratos", f"id=eq.{row['id']}", {"status": es, "adiantamento_pago": ea, "obs": eo}):
                            st.success("✅ Salvo!"); st.cache_data.clear(); st.rerun()
            if st.button(f"🗑️ Excluir {row['contrato']}", type="secondary"):
                if sb_delete("contratos", f"id=eq.{row['id']}"):
                    st.success("Excluído!"); st.cache_data.clear(); st.rerun()
    else:
        st.info("Nenhum contrato encontrado.")

# ══════════════════════════════════════════════════════════════════
# POR MOTORISTA
# ══════════════════════════════════════════════════════════════════
elif aba == "motorista":
    st.markdown("<h1>👤 Por Motorista</h1>", unsafe_allow_html=True)

    # Anos até 2030 + anos existentes no banco
    anos_banco_m = set(df_all["data"].dt.year.dropna().unique().astype(int).tolist()) if not df_all.empty else set()
    anos_list_m = [0] + sorted(anos_banco_m | set(range(2024, 2031)), reverse=True)

    # Filtros na mesma linha: Motorista | Ano | Mês
    cf1, cf2, cf3 = st.columns([2, 1, 1])
    with cf1:
        mot = st.selectbox("👤 Motorista", MOTORISTAS, key="mot_nome")
    with cf2:
        ano_m = st.selectbox("📅 Ano", anos_list_m,
            index=anos_list_m.index(datetime.now().year) if datetime.now().year in anos_list_m else 1,
            format_func=lambda x: "Todos" if x == 0 else str(x),
            key="mot_ano")
    with cf3:
        mes_m = st.selectbox("🗓️ Mês", [0] + list(range(1, 13)),
            index=datetime.now().month,
            format_func=lambda x: "Todos" if x == 0 else MESES[x - 1],
            key="mot_mes")

    # Filtrar dados pelo período e motorista
    df_m = df_all.copy()
    if not df_all.empty:
        if ano_m: df_m = df_m[df_m["data"].dt.year == ano_m]
        if mes_m: df_m = df_m[df_m["data"].dt.month == mes_m]
    dm = df_m[df_m["motorista"] == mot]

    periodo_m = f"{MESES[mes_m-1]}/{ano_m}" if mes_m and ano_m else \
                (MESES[mes_m-1] if mes_m else (str(ano_m) if ano_m else "Todos os períodos"))

    st.markdown(f"<p style='color:#8896A7;margin-top:-4px;font-size:13px'>📅 {periodo_m} · {len(dm)} contratos</p>", unsafe_allow_html=True)

    fat = dm["fat_bruto"].sum()
    a, f = com(mot, fat)
    pct = min(fat / META * 100, 100)
    cor = "#2ECC71" if fat >= META else "#E74C3C"

    # Card perfil
    st.markdown(f"""
    <div class='panel-card' style='margin-bottom:20px'>
        <div style='display:flex;justify-content:space-between;align-items:flex-start'>
            <div>
                <div style='font-size:24px;font-weight:800;color:#1A3A5C'>{mot}</div>
                <div style='font-size:12px;color:#8896A7;margin-top:4px'>{"Sem adiantamento · 10% folha" if mot in SEM else "Com adiantamento · 5% + 5%"}</div>
            </div>
            <div style='text-align:right'>
                <div style='font-size:28px;font-weight:800;color:{cor}'>{R_short(fat)}</div>
                <div style='font-size:12px;color:{"#2ECC71" if fat >= META else "#E74C3C"}'>{"🏆 Meta atingida!" if fat >= META else f"Falta {R_short(META - fat)}"}</div>
            </div>
        </div>
        <div style='margin-top:16px;background:#F4F6F9;border-radius:6px;height:10px'>
            <div style='background:{cor};width:{pct}%;height:10px;border-radius:6px;transition:width 0.5s'></div>
        </div>
        <div style='display:flex;justify-content:space-between;margin-top:6px'>
            <span style='font-size:11px;color:#8896A7'>R$ 0</span>
            <span style='font-size:11px;color:#F59E0B;font-weight:600'>Meta: {R_short(META)}</span>
            <span style='font-size:11px;color:#8896A7'>{pct:.0f}%</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Métricas
    c1, c2, c3, c4 = st.columns(4)
    for col, label, val, cor_m in [
        (c1, "📄 Viagens", str(len(dm)), "#1A7FC1"),
        (c2, "⏩ Adiantamento", R_short(a) if mot not in SEM else "N/A", "#F59E0B"),
        (c3, "📋 Folha", R_short(f), "#8B5CF6"),
        (c4, "🔧 Chapas", R_short(dm["chapa"].sum()), "#EC4899"),
    ]:
        with col:
            st.markdown(f"""<div class='metric-card' style='padding:14px 18px'>
                <div class='metric-label'>{label}</div>
                <div style='font-size:22px;font-weight:800;color:{cor_m}'>{val}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    if dm.empty:
        st.info(f"Nenhum contrato para **{mot}** em {periodo_m}.")
    else:
        cols = [c for c in ["contrato","cliente","data","fat_bruto","chapa","destino","status"] if c in dm.columns]
        ds = dm[cols].copy()
        ds["data"] = ds["data"].dt.strftime("%d/%m/%Y")
        ds["fat_bruto"] = ds["fat_bruto"].apply(R)
        ds["chapa"] = ds["chapa"].apply(R)
        ds.columns = [c.upper().replace("_"," ") for c in ds.columns]
        st.dataframe(ds, use_container_width=True, hide_index=True)
        try:
            excel_buf = gerar_excel(dm, f"{mot} {periodo_m}")
            st.download_button("📊 Exportar Excel", excel_buf,
                f"{mot.replace(' ','_')}_{periodo_m.replace('/','_')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except: pass

elif aba == "comissoes":
    st.markdown("<h1>💳 Comissões</h1>", unsafe_allow_html=True)
    df, periodo = sel_periodo("com")
    st.markdown(f"<p style='color:#8896A7;margin-top:-8px'>📅 {periodo}</p>", unsafe_allow_html=True)
    if df.empty:
        st.info("Nenhum dado no período.")
    else:
        for mot in MOTORISTAS:
            dm = df[df["motorista"] == mot]
            if dm.empty: continue
            fat = dm["fat_bruto"].sum()
            a, f = com(mot, fat)
            pct = min(fat / META * 100, 100)
            cor = "#2ECC71" if fat >= META else "#1A7FC1"
            st.markdown(f"""
            <div class='panel-card' style='margin-bottom:10px'>
                <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:10px'>
                    <div>
                        <span style='font-weight:700;color:#1A3A5C;font-size:14px'>{"🏆 " if fat >= META else ""}{mot}</span>
                        <span style='font-size:11px;color:#8896A7;margin-left:8px'>{"10% folha" if mot in SEM else "5%+5%"} · {len(dm)} viagens</span>
                    </div>
                    <span style='font-size:16px;font-weight:800;color:{cor}'>{R_short(fat)}</span>
                </div>
                <div style='display:flex;gap:24px;font-size:12px;color:#5A6A7A;flex-wrap:wrap;margin-bottom:10px'>
                    {"<span>Adiant: <b style='color:#F59E0B'>N/A</b></span>" if mot in SEM else f"<span>Adiant: <b style='color:#F59E0B'>{R(a)}</b></span>"}
                    <span>Folha: <b style='color:#8B5CF6'>{R(f)}</b></span>
                    <span>Chapas: <b style='color:#EC4899'>{R(dm["chapa"].sum())}</b></span>
                    <span>Total comissão: <b style='color:#2ECC71'>{R(a+f)}</b></span>
                </div>
                <div style='background:#F4F6F9;border-radius:4px;height:5px'>
                    <div style='background:{cor};width:{pct}%;height:5px;border-radius:4px'></div>
                </div>
            </div>""", unsafe_allow_html=True)

        st.markdown("---")
        rows_exp = []
        for mot in MOTORISTAS:
            dm = df[df["motorista"] == mot]
            if dm.empty: continue
            fat = dm["fat_bruto"].sum()
            a, f = com(mot, fat)
            rows_exp.append({"motorista": mot, "tipo": "10% folha" if mot in SEM else "5%+5%",
                "viagens": len(dm), "fat_bruto": fat, "adiantamento": 0 if mot in SEM else a,
                "folha": f, "chapas": dm["chapa"].sum(), "total_comissao": a+f})
        df_com = pd.DataFrame(rows_exp)
        try:
            excel_buf = gerar_excel(df_com, f"Comissões {periodo}")
            st.download_button("📊 Exportar Comissões (.xlsx)", excel_buf,
                f"comissoes_{periodo.replace('/','_')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except: pass

# ══════════════════════════════════════════════════════════════════
# PRÊMIOS
# ══════════════════════════════════════════════════════════════════
elif aba == "premios":
    st.markdown(f"<h1>🏆 Prêmios</h1>", unsafe_allow_html=True)
    df, periodo = sel_periodo("prem")
    st.markdown(f"<p style='color:#8896A7;margin-top:-8px'>Meta: {R(META)} · 📅 {periodo}</p>", unsafe_allow_html=True)
    if df.empty:
        st.info("Nenhum dado no período.")
    else:
        for mot in MOTORISTAS:
            dm = df[df["motorista"] == mot]
            if dm.empty: continue
            fat  = dm["fat_bruto"].sum()
            el   = fat >= META
            prem = prem_map.get(mot, {})
            with st.expander(f"{'🏆' if el else '⏳'} {mot} — {R_short(fat)}", expanded=el):
                c1,c2,c3 = st.columns(3)
                c1.metric("Faturamento", R_short(fat))
                c2.metric("Meta", R_short(META))
                c3.metric("Status", "✅ Elegível" if el else f"{min(fat/META*100,100):.0f}%")
                if prem.get("status"):
                    cor_p = {"QUALIFICADO":"#1A7FC1","EM ANÁLISE":"#F59E0B","APROVADO":"#2ECC71","PAGO":"#10B981","NÃO APROVADO":"#E74C3C"}.get(prem["status"],"#8896A7")
                    st.markdown(f"""<div style='background:#F8FAFC;border-left:4px solid {cor_p};border-radius:0 8px 8px 0;padding:10px 14px;margin:8px 0'>
                        <span style='color:{cor_p};font-weight:700'>{prem["status"]}</span>
                        {f" · <b style='color:#2ECC71'>{R(prem.get('valor',0))}</b>" if prem.get("valor") else ""}
                        {f"<div style='color:#5A6A7A;font-size:12px;margin-top:4px'>{prem['obs']}</div>" if prem.get("obs") else ""}
                    </div>""", unsafe_allow_html=True)
                if perm == "total" and el:
                    with st.form(f"fp_{mot}"):
                        pc1,pc2 = st.columns(2)
                        ps = pc1.selectbox("Status", STATUS_P,
                            index=STATUS_P.index(prem.get("status","QUALIFICADO")) if prem.get("status") in STATUS_P else 0)
                        pv = pc2.number_input("Valor (R$)", value=float(prem.get("valor") or 0), format="%.2f")
                        po = st.text_input("Observação", value=prem.get("obs","") or "")
                        if st.form_submit_button("💾 Salvar Prêmio"):
                            if sb_post("premios", {"motorista": mot, "status": ps, "valor": pv, "obs": po}, upsert=True):
                                st.success("✅ Salvo!"); st.cache_data.clear(); st.rerun()
